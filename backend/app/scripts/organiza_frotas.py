import pandas as pd
import os
from datetime import datetime
import openpyxl
import sys
import locale

# Configurar locale para pt_BR para tratar números corretamente
# Ajustar configuração do locale para lidar com falhas
try:
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
except locale.Error:
    try:
        locale.setlocale(locale.LC_ALL, 'Portuguese_Brazil.1252')
    except locale.Error:
        print("Aviso: Não foi possível configurar o locale para pt_BR. Usando configurações padrão.")

def converter_numero(valor):
    """Converte strings de números para float, tratando diferentes formatos"""
    if pd.isna(valor):
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    try:
        # Remove espaços e substitui vírgula por ponto
        valor_limpo = str(valor).strip().replace('.', '').replace(',', '.')
        return float(valor_limpo)
    except:
        return 0.0

def estilizar_celula_botao(cell):
    """Aplica estilo de botão a uma célula"""
    cell.font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    cell.border = Border(
        left=Side(style='thick', color='000000'),
        right=Side(style='thick', color='000000'),
        top=Side(style='thick', color='000000'),
        bottom=Side(style='thick', color='000000')
    )
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Permitir argumentos: custos, frotas, saida
if len(sys.argv) >= 3:
    caminho_custos = os.path.abspath(sys.argv[1])
    caminho_frotas = os.path.abspath(sys.argv[2])
    caminho_saida = os.path.abspath(sys.argv[3] if len(sys.argv) > 3 else 'planilha_organizada.xlsx')
else:
    caminho_custos = 'CUSTO 01-01 A 31-05 teste.xlsx'
    caminho_frotas = 'Relação de frotas.xlsx'
    caminho_saida = 'planilha_organizada.xlsx'

# Verificar se os arquivos existem
for caminho, desc in [(caminho_custos, 'custos'), (caminho_frotas, 'frotas')]:
    if not os.path.exists(caminho):
        print(f"Erro: Arquivo de {desc} não encontrado: {caminho}")
        sys.exit(1)

# Garantir que o diretório de saída existe
diretorio_saida = os.path.dirname(caminho_saida)
if diretorio_saida and not os.path.exists(diretorio_saida):
    try:
        os.makedirs(diretorio_saida)
    except Exception as e:
        print(f"Erro ao criar diretório de saída: {str(e)}")
        sys.exit(1)

# Leitura das planilhas com tratamento de números
custos = pd.read_excel(caminho_custos, dtype={'NUM_FROTA': str})
frotas = pd.read_excel(caminho_frotas, dtype={'NUM_FROTA': str})

    # Converte colunas numéricas
for col in ['QTDE_ITEM', 'VLR_TOT_ITEM', 'VLR_ICMS']:
    if col in custos.columns:
        custos[col] = custos[col].apply(converter_numero)

try:
    # Análise inicial das planilhas
    print('--- ANÁLISE PLANILHA DE CUSTOS ---')
    print('Colunas:', list(custos.columns))
    print(custos.head())
    print('\n--- ANÁLISE PLANILHA DE FROTAS ---')
    print('Colunas:', list(frotas.columns))
    print(frotas.head())

    # Normaliza datas e extrai mês
    if 'DTA_MOVIMENTO' in custos.columns:
        custos['Mês'] = pd.to_datetime(custos['DTA_MOVIMENTO']).dt.strftime('%m/%Y')
    else:
        custos['Mês'] = ''

    # Junta custos com dados da frota
    dados = pd.merge(custos, frotas, on='NUM_FROTA', how='left')
except Exception as e:
    print(f"Erro durante a análise inicial das planilhas: {str(e)}")
    sys.exit(1)

try:
    # Verifica se existe a coluna ANO_MODELO_VEICULO na tabela de frotas
    if 'ANO_MODELO_VEICULO' not in frotas.columns:
        # Se não existir, criar a coluna com valores vazios
        frotas['ANO_MODELO_VEICULO'] = ''
        print("Coluna ANO_MODELO_VEICULO não encontrada na planilha de frotas. Criando coluna vazia.")

    # Pivot para organizar por tipo de gasto
    if 'DSC_ITEMTABELA' in dados.columns:
        tabela = dados.pivot_table(index=['Mês', 'NUM_FROTA', 'ANO_MODELO_VEICULO', 'DSC_MARCA', 'DSC_TIPO'],
                                columns='DSC_ITEMTABELA',
                                values='VLR_TOT_ITEM',
                                aggfunc='sum',
                                fill_value=0).reset_index()
    else:
        tabela = dados

    # Renomeia colunas para facilitar
    colunas_renomeadas = {
        'Combustíveis e Lubrificantes': 'Combustível',
        'Pedágio': 'Pedágios',
        'Viagens': 'Viagens e Estadias',
        # Adicione outros tipos de gasto conforme necessário
    }
    tabela = tabela.rename(columns=colunas_renomeadas)
except Exception as e:
    print(f"Erro durante o processamento dos dados: {str(e)}")
    sys.exit(1)

try:
    # Calcula o total de despesas
    # Identifica colunas numéricas (excluindo colunas de identificação como Mês, NUM_FROTA, etc)
    colunas_despesas = [col for col in tabela.columns if col not in ['Mês', 'NUM_FROTA', 'ANO_MODELO_VEICULO', 'DSC_MARCA', 'DSC_TIPO']]
    tabela['Total Despesas'] = tabela[colunas_despesas].sum(axis=1)

    # Reordena as colunas para colocar Total Despesas após Viagens e Estadias
    todas_colunas = list(tabela.columns)
    if 'Viagens e Estadias' in todas_colunas:
        # Remove Total Despesas da lista
        if 'Total Despesas' in todas_colunas:
            todas_colunas.remove('Total Despesas')
        # Encontra a posição da coluna Viagens e Estadias
        pos_viagens = todas_colunas.index('Viagens e Estadias')
        # Insere Total Despesas logo após Viagens e Estadias
        todas_colunas.insert(pos_viagens + 1, 'Total Despesas')
        # Reordena a tabela
        tabela = tabela[todas_colunas]

    # Salva a planilha organizada com estilização
    with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
        tabela.to_excel(writer, index=False)
        
        # Acessa o arquivo para formatação
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
except Exception as e:
    print(f"Erro durante o processamento dos totais e reordenação: {str(e)}")
    sys.exit(1)

try:
    # Configurações de estilo
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Color
    
    # Adicionar uma linha no topo para o título
    worksheet.insert_rows(1)
    
    # Configurar o título
    titulo_range = f'A1:{openpyxl.utils.get_column_letter(worksheet.max_column)}1'
    worksheet.merge_cells(titulo_range)
    titulo_cell = worksheet['A1']
    titulo_cell.value = "DEMONSTRATIVO DE FECHAMENTO FERTFER"
    titulo_cell.font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
    titulo_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajustar altura da linha do título
    worksheet.row_dimensions[1].height = 30
    
    # Estilos para cabeçalho de dados
    header_font = Font(bold=True, size=11, color="FFFFFF")  # Texto branco
    header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Fundo vermelho
    
    # Definindo bordas mais espessas
    thick_border = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Aplicar estilo no cabeçalho
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thick_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Fonte padrão para todas as células
    default_font = Font(name='Arial', size=10)
    
    # Formatar todas as células com borda
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = default_font
            
    # Estilo especial para a última linha (totais)
    for cell in worksheet[worksheet.max_row]:
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.border = thick_border

    # Ajustar largura das colunas com tamanhos mínimos
    min_widths = {
        'Mês': 12,
        'NUM_FROTA': 10,
        'DSC_MARCA': 15,
        'DSC_TIPO': 15,
        'ANO_MODELO_VEICULO': 10,
        'Total Despesas': 15
    }
    
    # Função auxiliar para verificar se uma célula é mesclada
    def is_merged_cell(cell):
        return isinstance(cell, openpyxl.cell.cell.MergedCell)

    # Ajustar formatação para evitar erro com células mescladas
    for column in worksheet.columns:
        col_letter = column[0].column_letter if not is_merged_cell(column[0]) else None
        col_name = worksheet.cell(row=1, column=column[0].column).value if col_letter else None

        # Calcula a largura máxima do conteúdo
        max_length = 0
        for cell in column:
            if not is_merged_cell(cell):
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

        # Usa o valor mínimo definido ou calculado
        min_width = min_widths.get(col_name, 10) if col_name else 10
        adjusted_width = max(min_width, max_length + 2)

        if col_letter:
            worksheet.column_dimensions[col_letter].width = adjusted_width
    
    try:
        # Criar botão Menu no topo da planilha
        menu_cell = 'Q1'  # Posição do botão Menu
        worksheet[menu_cell] = "Menu"
        estilizar_celula_botao(worksheet[menu_cell])
        worksheet.column_dimensions[menu_cell[0]].width = 15  # Largura da coluna do Menu

        # Adicionar outros botões
        botoes = [
            ('R1', 'RODOTREM'),
            ('S1', 'TRUCK'),
            ('T1', 'CAVALO RESERVA'),
            ('U1', 'GRANELEIRA'),
            ('V1', 'BI-CAÇAMBA'),
            ('W1', 'ÍNDICES')
        ]

        for pos, texto in botoes:
            worksheet[pos] = texto
            estilizar_celula_botao(worksheet[pos])
            worksheet.column_dimensions[pos[0]].width = max(15, len(texto) + 2)
    except Exception as e:
        print(f"Aviso: Não foi possível criar todos os botões: {str(e)}")
        # Continua a execução mesmo se houver erro nos botões
        
    # Congelar painel no cabeçalho
    worksheet.freeze_panes = 'A2'

    # Aplicar formatação específica para cada tipo de coluna
    for idx, col_name in enumerate(tabela.columns, 1):
        col_letter = worksheet.cell(row=1, column=idx).column_letter
        
        # Colunas monetárias
        if any(word in col_name for word in ['Despesas', 'Combustível', 'Pedágios', 'Viagens', 'Comb.', 'Lub.']):
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet[f"{col_letter}{row}"]
                cell.number_format = 'R$ #,##0.00'
                # Alinhamento à direita para valores monetários
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if col_name == 'Total Despesas':
                    cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
                    cell.font = Font(name='Arial', size=10, bold=True)
        
        # Colunas de quilometragem ou números inteiros
        elif any(term in col_name for term in ['KM', 'Qtd', 'Quantidade']):
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet[f"{col_letter}{row}"]
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
                
        # Colunas de texto (como Mês, NUM_FROTA, etc)
        else:
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet[f"{col_letter}{row}"]
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Adicionar filtros no cabeçalho
    worksheet.auto_filter.ref = worksheet.dimensions
except Exception as e:
    print(f"Erro durante a formatação da planilha: {str(e)}")
    sys.exit(1)

print(f'Planilha organizada salva em: {os.path.abspath(caminho_saida)}')

# Verifica se a planilha destino existe antes de tentar atualizá-la
caminho_planilha_destino = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'JANEIRO A DEZEMBRO - CUSTOS 2025 copia.xlsm')

if os.path.exists(caminho_planilha_destino):
    print(f'Analisando e atualizando a planilha destino: {caminho_planilha_destino}')
    wb = openpyxl.load_workbook(caminho_planilha_destino, keep_vba=True)
    resumo_alteracoes = []

    # Mapeamento dos nomes das colunas conforme a imagem
    colunas_gastos = {
        'Combustível': 'Comb. e Lub.',
        'Veículo': 'Gastos c/Veículos',
        'Pedágios': 'Pedágio / Rastreador',
        'Lanches/Diárias': 'Ref. e Lanches (Diárias)',
        'Viagens e Estadias': 'Viagens e Estadias',
        'Total Despesas': 'Total Despesas',
    }

    for aba in wb.sheetnames:
        ws = wb[aba]
        for _, row_org in tabela.iterrows():
            num_frota = row_org['NUM_FROTA']
            # Procura linha da frota
            linha_frota = None
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if str(row[0].value) == str(num_frota):
                    linha_frota = row[0].row
                    break
            if linha_frota:
                # Atualiza valores apenas nas colunas de gastos
                for col_org, col_dest in colunas_gastos.items():
                    valor = row_org.get(col_org, 0)
                    coluna_dest = None
                    for cell in ws[1]:
                        if cell.value and col_dest.lower() in str(cell.value).lower():
                            coluna_dest = cell.column
                            break
                    if coluna_dest:
                        ws.cell(row=linha_frota, column=coluna_dest, value=valor)
                        resumo_alteracoes.append(f'Aba: {aba}, Frota: {num_frota}, Coluna: {col_dest}, Valor: {valor}')

    wb.save(caminho_planilha_destino)
    print('Resumo das alterações realizadas:')
    for alt in resumo_alteracoes:
        print(alt)
    print('Planilha destino atualizada com segurança!')

    # Análise dos valores inseridos na planilha destino
    print('\n--- ANÁLISE DOS VALORES INSERIDOS NA PLANILHA DESTINO ---')
    for aba in wb.sheetnames:
        ws = wb[aba]
        print(f'\nAba: {aba}')
        # Exibe os valores das colunas de gastos para cada frota
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            num_frota = row[0].value
            if num_frota:
                valores = {}
                for col_dest in colunas_gastos.values():
                    coluna_idx = None
                    for cell in ws[1]:
                        if cell.value and col_dest.lower() in str(cell.value).lower():
                            coluna_idx = cell.column - 1
                            break
                    if coluna_idx is not None:
                        valores[col_dest] = row[coluna_idx].value
                print(f'Frota: {num_frota} | ' + ' | '.join([f'{col}: {valores[col]}' for col in valores]))
else:
    print(f"Arquivo '{caminho_planilha_destino}' não encontrado. Apenas a planilha organizada será gerada.")
